import copy

from openpyxl import load_workbook


class SystemImport:
    def __init__(self, excelfiles):
        wb = load_workbook(excelfiles, read_only=True, data_only=True)
        print(wb.sheetnames)
        datatypes_sheet = wb['DataTypes']
        interfaces_sheet = wb['Interfaces']
        self.collect_datatypes(datatypes_sheet)
        self.collect_portinterfaces(interfaces_sheet)

    def collect_portinterfaces(self, worksheet):
        self.interfaceslist = []
        interface = {}
        for row in worksheet:
            # print(row[1].value)
            if "S/R" == row[2].value:
                # print(row[2].value)
                interface = \
                {
                    'name': row[1].value,
                    'type': row[2].value,
                    'elements': [],
                    'element type': []
                }
            if {} != interface:
                if row[3].value is not None:
                    interface['elements'].append(row[3].value)
                    interface['element type'].append(row[4].value)
                else:
                    self.interfaceslist.append(interface)
                    interface = {}

    def collect_datatypes(self, worksheet):
        self.basetypelist = []
        self.enumtypelist = []
        self.structtypelist = []
        basetype = {}
        enumtype = {}
        structtype = {}
        for row in worksheet:
            if "Base Types" == row[2].value:
                basetype['name'] = row[1].value
                basetype['bit size'] = row[3].value
                basetype['native declaration'] = row[4].value
                self.basetypelist.append(basetype)
                basetype = {}
            if "Enum" == row[2].value:
                enumtype = \
                    {
                        'name': row[1].value,
                        'value table': [],
                        'value': []
                    }
            if {} != enumtype:
                if row[3].value is not None:
                    enumtype['value table'].append(row[3].value)
                    enumtype['value'].append(row[4].value)
                else:
                    self.enumtypelist.append(enumtype)
                    enumtype = {}
            if "Struct" == row[2].value:
                structtype = \
                    {
                        'name': row[1].value,
                        'elements': [],
                        'element type': []
                    }
            if {} != structtype:
                if row[3].value is not None:
                    structtype['elements'].append(row[3].value)
                    structtype['element type'].append(row[4].value)
                else:
                    self.structtypelist.append(structtype)
                    structtype = {}


if __name__ == '__main__':
    System = SystemImport('Application.xlsx')
    # print(System.basetypelist)
    # print(System.enumtypelist)
    # print(System.structtypelist)
    print(System.interfaceslist)
