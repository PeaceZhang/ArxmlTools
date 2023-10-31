import copy
from openpyxl import load_workbook


class SystemImport:
    def __init__(self, excelfiles):
        wb = load_workbook(excelfiles, read_only=True, data_only=True)
        # print(wb.sheetnames)
        datatypes_sheet = wb['DataTypes']
        interfaces_sheet = wb['Interfaces']
        component_sheet = wb['SWC_Composition']
        self.collect_datatypes(datatypes_sheet)
        self.collect_portinterfaces(interfaces_sheet)
        self.collect_swcomponent(component_sheet)
    def collect_swcomponent(self, worksheet):
        self.swcomponentlist = []
        swcomponent = {}
        for row in worksheet:
            if row[2].value not in ['ComponentType', None]:
                swcomponent = \
                    {
                        'swc name': row[1].value,
                        'componenttype': row[2].value,
                        'elements': []
                    }
            if {} != swcomponent:
                if row[3].value is not None:
                    # swcomponent['componenttype'].append(row[2].value)
                    attribute = {}
                    attribute['attributes'] = row[3].value
                    attribute['portname'] = row[4].value
                    attribute['direction'] = row[5].value
                    attribute['defaultportaccess'] = row[6].value
                    attribute['schedule'] = row[7].value
                    attribute['interfaces'] = row[8].value
                    attribute['receiver swc'] = row[9].value
                    if None != row[9].value:
                        attribute['receiver swc'] = [x.strip() for x in row[9].value.strip().split(",")]
                    swcomponent['elements'].append(attribute)
                else:
                    self.swcomponentlist.append(swcomponent)
                    swcomponent = {}

    def collect_portinterfaces(self, worksheet):
        self.interfaceslist = []
        self.csinterfaceslist = []
        interface = {}
        csinterface = {}
        operation = {}
        argument = []
        for row in worksheet:
            # print(row[1].value)
            if "S/R" == row[2].value:
                # print(row[2].value)
                interface = \
                {
                    'name': row[1].value,
                    'type': row[2].value,
                    'elements': [],
                    'element type': []
                }
            if {} != interface:
                if row[3].value is not None:
                    interface['elements'].append(row[3].value)
                    interface['element type'].append(row[4].value)
                else:
                    self.interfaceslist.append(interface)
                    interface = {}

            # collect cs interfaces list
            if "C/S" == row[2].value:
                csinterface = \
                    {
                        'name': row[1].value,
                        'type': row[2].value,
                        'application error': row[3].value,
                        'operations': []
                    }
            if {} != csinterface:
                if None != row[4].value or None != row[5].value:
                    if None != row[4].value:
                        if {} != operation:
                            csinterface['operations'].append(operation)
                            operation = {}
                        # create new operation
                        operation = \
                        {
                            'operation name': row[4].value,
                            'arguments': []
                        }
                    if None != row[5].value:
                        argument = \
                            {
                                'argument name': row[5].value,
                                'argstype': row[6].value,
                                'direction': row[7].value
                            }
                        operation['arguments'].append(argument)
                else:
                    csinterface['operations'].append(operation)
                    self.csinterfaceslist.append(csinterface)
                    csinterface = {}
                    operation = {}
                    argument = []

    def collect_datatypes(self, worksheet):
        self.basetypelist = []
        self.aliastypelist = []
        self.enumtypelist = []
        self.enumaetypelist = []
        self.structtypelist = []
        basetype = {}
        aliastype = {}
        enumtype = {}
        enumaetype = {}
        structtype = {}
        for row in worksheet:
            # collect base type list
            if "Base Types" == row[2].value:
                basetype['name'] = row[1].value
                basetype['bit size'] = row[3].value
                basetype['native declaration'] = row[4].value
                self.basetypelist.append(basetype)
                basetype = {}

            # collect alias type list
            if "Alias" == row[2].value:
                aliastype['name'] = row[1].value
                aliastype['reference type'] = row[3].value
                aliastype['scaling'] = row[4].value
                aliastype['offset'] = row[5].value
                aliastype['minvalue'] = row[6].value
                aliastype['maxvalue'] = row[7].value
                aliastype['unit'] = row[8].value
                self.aliastypelist.append(aliastype)
                aliastype = {}

            # collect enum type list
            if "Enum" == row[2].value:
                enumtype = \
                    {
                        'name': row[1].value,
                        'value table': [],
                        'value': []
                    }
            if {} != enumtype:
                if row[3].value is not None:
                    enumtype['value table'].append(row[3].value)
                    enumtype['value'].append(row[4].value)
                else:
                    self.enumtypelist.append(enumtype)
                    enumtype = {}

            # collect application possible error list
            if "EnumAppError" == row[2].value:
                enumaetype = \
                    {
                        'name': row[1].value,
                        'value table': [],
                        'value': []
                    }
            if {} != enumaetype:
                if row[3].value is not None:
                    enumaetype['value table'].append(row[3].value)
                    enumaetype['value'].append(row[4].value)
                else:
                    self.enumaetypelist.append(enumaetype)
                    enumaetype = {}

            # collect struct type list
            if "Struct" == row[2].value:
                structtype = \
                    {
                        'name': row[1].value,
                        'elements': [],
                        'element type': []
                    }
            if {} != structtype:
                if row[3].value is not None:
                    structtype['elements'].append(row[3].value)
                    structtype['element type'].append(row[4].value)
                else:
                    self.structtypelist.append(structtype)
                    structtype = {}


if __name__ == '__main__':
    System = SystemImport('../Import/Application.xlsx')
    # print(System.basetypelist)
    # print(System.enumtypelist)
    # print(System.enumaetypelist)
    # print(System.structtypelist)
    # print(System.interfaceslist)
    # print(System.swcomponentlist)
    # print(System.aliastypelist)
    print(System.csinterfaceslist)

